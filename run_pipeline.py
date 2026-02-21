# run_pipeline.py
# Minimal, repo-friendly runner that works in terminal AND in notebooks (Colab/Jupyter).

import os
import re
from datetime import datetime

import numpy as np
import pandas as pd

import matplotlib
matplotlib.use("Agg")  # safe for headless runs (e.g., Codespaces/CI)
import matplotlib.pyplot as plt

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as XLImage


# -------------------------
# Helpers
# -------------------------

def normalize_colname(c: str) -> str:
    return re.sub(r"\s+", " ", str(c).strip())

def find_col(cols, candidates):
    norm_map = {normalize_colname(c).lower(): c for c in cols}
    for cand in candidates:
        key = normalize_colname(cand).lower()
        if key in norm_map:
            return norm_map[key]

    stripped_map = {re.sub(r"\s+", "", normalize_colname(c)).lower(): c for c in cols}
    for cand in candidates:
        key = re.sub(r"\s+", "", normalize_colname(cand)).lower()
        if key in stripped_map:
            return stripped_map[key]

    return None

def is_int_like(x):
    if pd.isna(x):
        return False
    if isinstance(x, (int, np.integer)):
        return True
    if isinstance(x, (float, np.floating)):
        return float(x).is_integer()
    if isinstance(x, str):
        return x.strip().isdigit()
    return False

def to_int(x):
    if isinstance(x, (int, np.integer)):
        return int(x)
    if isinstance(x, (float, np.floating)):
        return int(round(float(x)))
    if isinstance(x, str) and x.strip().isdigit():
        return int(x.strip())
    return None

def interpret_match_value(v):
    """Return True for Match, False for Mismatch, np.nan for unknown."""
    if pd.isna(v):
        return np.nan
    if isinstance(v, (bool, np.bool_)):
        return bool(v)

    if isinstance(v, (int, np.integer)):
        return bool(v) if v in (0, 1) else np.nan

    if isinstance(v, (float, np.floating)):
        if float(v).is_integer() and int(v) in (0, 1):
            return bool(int(v))
        return np.nan

    s = str(v).strip().lower()
    if s in ("match", "matched", "m", "true", "t", "yes", "y", "1"):
        return True
    if s in ("mismatch", "not match", "notmatch", "mm", "false", "f", "no", "n", "0"):
        return False
    return np.nan

def valid_indicator_series(ser: pd.Series) -> bool:
    vals = ser.dropna().map(interpret_match_value)
    return vals.notna().all()

def write_df_to_sheet(ws, df: pd.DataFrame, index=False, header=True):
    ws.delete_rows(1, ws.max_row)
    for r_idx, row in enumerate(dataframe_to_rows(df, index=index, header=header), start=1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

def save_fig(path, fig):
    fig.savefig(path, dpi=200, bbox_inches="tight")
    plt.close(fig)


# -------------------------
# Core pipeline
# -------------------------

def run_pipeline(src_path: str, out_path: str) -> str:
    # Detect input sheets
    wb_in = load_workbook(src_path)

    sheet_dup = None
    for cand in ["DuplicatePatientsSummary", "DuplicatePatientSummary", "DuplicatePatients", "DuplicatePatients Summary"]:
        if cand in wb_in.sheetnames:
            sheet_dup = cand
            break
    if sheet_dup is None:
        raise ValueError(f"Could not find DuplicatePatientsSummary-like sheet. Found: {wb_in.sheetnames}")

    sheet_input = None
    for cand in ["Input_DuplicatePatientSummary", "Input_DuplicatePatientsSummary", "Input DuplicatePatientSummary"]:
        if cand in wb_in.sheetnames:
            sheet_input = cand
            break

    # Column mapping
    cols_dup = pd.read_excel(src_path, sheet_name=sheet_dup, nrows=0, engine="openpyxl").columns.tolist()

    col_patient1 = find_col(cols_dup, ["Patient 1", "Patient1"])
    col_patient2 = find_col(cols_dup, ["Patient 2", "Patient2"])
    col_score    = find_col(cols_dup, ["MatchScore", "Match Score", "Score"])

    col_forename = find_col(cols_dup, ["Forename", "Forename Match", "Forename_Match"])
    col_surname  = find_col(cols_dup, ["Surname", "Surname Match", "Surname_Match"])
    col_soundex  = find_col(cols_dup, ["Soundex", "Soundex Match", "Soundex_Match"])
