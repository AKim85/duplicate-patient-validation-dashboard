# modules/M1_schema_validation.py
# Module 1: Data Readiness (Schema Validation)
# Writes results to Excel tab: M1_SchemaValidation
# Returns: (df, col_map, halt, schema_results, sheet_used)

import re
import numpy as np
import pandas as pd

from openpyxl.utils.dataframe import dataframe_to_rows


# -------------------------
# Helpers
# -------------------------

def _normalize_colname(c: str) -> str:
    return re.sub(r"\s+", " ", str(c).strip())

def _find_col(cols, candidates):
    """Case-insensitive, whitespace-tolerant column lookup."""
    norm_map = {_normalize_colname(c).lower(): c for c in cols}
    for cand in candidates:
        key = _normalize_colname(cand).lower()
        if key in norm_map:
            return norm_map[key]

    # allow minor variants: remove spaces
    stripped_map = {re.sub(r"\s+", "", _normalize_colname(c)).lower(): c for c in cols}
    for cand in candidates:
        key = re.sub(r"\s+", "", _normalize_colname(cand)).lower()
        if key in stripped_map:
            return stripped_map[key]

    return None

def _is_int_like(x):
    if pd.isna(x):
        return False
    if isinstance(x, (int, np.integer)):
        return True
    if isinstance(x, (float, np.floating)):
        return float(x).is_integer()
    if isinstance(x, str):
        return x.strip().isdigit()
    return False

def _to_int(x):
    if isinstance(x, (int, np.integer)):
        return int(x)
    if isinstance(x, (float, np.floating)):
        return int(round(float(x)))
    if isinstance(x, str) and x.strip().isdigit():
        return int(x.strip())
    return None

def _interpret_match_value(v):
    """Return True for Match, False for Mismatch, np.nan for unknown."""
    if pd.isna(v):
        return np.nan
    if isinstance(v, (bool, np.bool_)):
        return bool(v)

    # numeric 0/1
    if isinstance(v, (int, np.integer)):
        if v in (0, 1):
            return bool(v)
        return np.nan
    if isinstance(v, (float, np.floating)):
        if float(v).is_integer() and int(v) in (0, 1):
            return bool(int(v))
        return np.nan

    s = str(v).strip().lower()
    if s in ("match", "matched", "m", "true", "t", "yes", "y", "1"):
        return True
    if s in ("mismatch", "not match", "notmatch", "mm", "false", "f", "no", "n", "0"):
        return False
    return np.nan

def _valid_indicator_series(ser: pd.Series) -> bool:
    vals = ser.dropna().map(_interpret_match_value)
    return vals.notna().all()

def _write_df_to_sheet(ws, df: pd.DataFrame, index=False, header=True):
    """Clear a worksheet and write a dataframe to it."""
    ws.delete_rows(1, ws.max_row)
    for r_idx, row in enumerate(dataframe_to_rows(df, index=index, header=header), start=1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)


# -------------------------
# Main entry
# -------------------------

def run(src_path: str, wb_out, sheet_name: str = None):
    """
    Validate schema for DuplicatePatientsSummary.
    Writes a Pass/Fail report to tab 'M1_SchemaValidation' and returns:
      df, col_map, halt, schema_results, sheet_used
    """

    # Detect sheet if not provided
    if sheet_name is None:
        candidates = ["DuplicatePatientsSummary", "DuplicatePatientSummary", "DuplicatePatients", "DuplicatePatients Summary"]
        sheet_name = next((s for s in candidates if s in wb_out.sheetnames), None)
        if sheet_name is None:
            # If wb_out came from copy of source, sheetnames should include it.
            raise ValueError(f"Could not find DuplicatePatientsSummary-like sheet. Found: {wb_out.sheetnames}")

    sheet_used = sheet_name

    # Read header row only
    cols = pd.read_excel(src_path, sheet_name=sheet_used, nrows=0, engine="openpyxl").columns.tolist()

    # Map required columns (tolerant to minor naming differences)
    col_patient1 = _find_col(cols, ["Patient 1", "Patient1"])
    col_patient2 = _find_col(cols, ["Patient 2", "Patient2"])
    col_score    = _find_col(cols, ["MatchScore", "Match Score", "Score"])

    col_forename = _find_col(cols, ["Forename", "Forename Match", "Forename_Match"])
    col_surname  = _find_col(cols, ["Surname", "Surname Match", "Surname_Match"])
    col_soundex  = _find_col(cols, ["Soundex", "Soundex Match", "Soundex_Match"])
    col_sex      = _find_col(cols, ["Sex", "Sex Match", "Sex_Match"])
    col_dob      = _find_col(cols, ["DOB", "Date of Birth", "DOB Match", "DOB_Match"])
    col_postcode = _find_col(cols, ["Poscode", "PostCode", "Post Code", "Postcode", "PostCode Match", "PostCode_Match", "Postcode Match"])
    col_address  = _find_col(cols, ["Address", "Address Match", "Address_Match"])

    col_map = {
        "Patient 1": col_patient1,
        "Patient 2": col_patient2,
        "MatchScore": col_score,
        "Forename": col_forename,
        "Surname": col_surname,
        "Soundex": col_soundex,
        "Sex": col_sex,
        "DOB": col_dob,
        "PostCode/Poscode": col_postcode,
        "Address": col_address,
    }

    # Build checks
    checks = []
    for k, v in col_map.items():
        checks.append({
            "Check": f"Column present: {k}",
            "Result": "Pass" if v is not None else "Fail",
            "Details": "" if v is not None else "Missing"
        })

    halt = any(c["Result"] == "Fail" for c in checks)

    # Read only available columns
    usecols = [c for c in col_map.values() if c is not None]
    df = pd.read_excel(src_path, sheet_name=sheet_used, engine="openpyxl", usecols=usecols)

    # MatchScore checks
    if col_score is not None:
        s = df[col_score]
        all_int_like = s.map(_is_int_like).all()
        in_range = s.map(lambda x: (_to_int(x) is not None) and (4 <= _to_int(x) <= 7)).all()

        checks.append({
            "Check": "'MatchScore' is integer-like",
            "Result": "Pass" if all_int_like else "Fail",
            "Details": "" if all_int_like else "Non-integer values present"
        })
        checks.append({
            "Check": "'MatchScore' within 4-7",
            "Result": "Pass" if in_range else "Fail",
            "Details": "" if in_range else "Out-of-range values present"
        })

        if (not all_int_like) or (not in_range):
            halt = True

    # Indicator columns checks
    indicator_cols = {
        "Forename": col_forename,
        "Surname": col_surname,
        "Soundex": col_soundex,
        "Sex": col_sex,
        "DOB": col_dob,
        "PostCode/Poscode": col_postcode,
        "Address": col_address,
    }

    for label, col in indicator_cols.items():
        if col is None:
            continue
        ok = _valid_indicator_series(df[col])
        checks.append({
            "Check": f"Indicator valid (Match/Mismatch or boolean/0-1): {label}",
            "Result": "Pass" if ok else "Fail",
            "Details": "" if ok else "Unexpected values present"
        })
        if not ok:
            halt = True

    schema_results = pd.DataFrame(checks)

    # Ensure destination tab exists
    if "M1_SchemaValidation" in wb_out.sheetnames:
        ws = wb_out["M1_SchemaValidation"]
    else:
        ws = wb_out.create_sheet("M1_SchemaValidation")

    _write_df_to_sheet(ws, schema_results, index=False)

    # Add a clear halt message at bottom if required
    if halt:
        ws.cell(row=schema_results.shape[0] + 3, column=1, value="Processing halted due to critical schema validation failures.")

    return df, col_map, halt, schema_results, sheet_used
