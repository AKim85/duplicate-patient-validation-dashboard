# modules/M4_triage_categorisation.py
import os
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils.dataframe import dataframe_to_rows

from modules.helpers import write_df_to_sheet, save_fig_to_png, contains_todelete

def run(df_std: pd.DataFrame, col_map: dict, wb_out,
        tab_name: str = "M4_TriageCategorisation", out_dir: str = "."):
    """
    Applies five-route triage logic and writes:
      - row-level triage list (includes Patient 1/2 IDs)
      - summary table (Count + %) at columns G:H:I
      - donut chart embedded on sheet
    Returns: triage_out_df, triage_summary_series, img_path
    """
    p1_col = col_map["Patient 1"]
    p2_col = col_map["Patient 2"]

    def triage_row(row):
        if contains_todelete(row):
            return "Delete"
        if row.get("DOB_is_match", np.nan) is False:
            return "Do-Not-Merge"

        sc = row.get("MatchScore_std", None)

        if sc == 7:
            return "Auto-Merge"

        if sc == 6 and row.get("Soundex_is_match", np.nan) is False:
            # Auto-merge only if Soundex is the only mismatch
            others = [
                "Forename_is_match", "Surname_is_match", "Sex_is_match",
                "DOB_is_match", "PostCode/Poscode_is_match", "Address_is_match"
            ]
            ok = True
            for f in others:
                if f in row.index and row.get(f) is False:
                    ok = False
                    break
            if ok:
                return "Auto-Merge"

        if row.get("Forename_is_match", np.nan) is False or row.get("Surname_is_match", np.nan) is False or row.get("Sex_is_match", np.nan) is False:
            return "Deep-Review"

        return "Fast-Track"

    triage_series = df_std.apply(triage_row, axis=1)

    triage_summary = triage_series.value_counts().reindex(
        ["Auto-Merge", "Fast-Track", "Deep-Review", "Do-Not-Merge", "Delete"],
        fill_value=0
    )

    triage_out = pd.DataFrame({
        "Patient 1": df_std[p1_col].astype(str),
        "Patient 2": df_std[p2_col].astype(str),
        "MatchScore": df_std["MatchScore_std"],
        "TriageRoute": triage_series
    })

    if tab_name in wb_out.sheetnames:
        ws = wb_out[tab_name]
    else:
        ws = wb_out.create_sheet(tab_name)

    write_df_to_sheet(ws, triage_out, index=False)

    # Summary table at column G
    summary_df = triage_summary.reset_index()
    summary_df.columns = ["TriageRoute", "Count"]
    total_pairs = int(triage_summary.sum())
    summary_df["Percentage"] = np.round((summary_df["Count"] / total_pairs) * 100, 2) if total_pairs else 0

    for r_idx, row in enumerate(dataframe_to_rows(summary_df, index=False, header=True), start=1):
        for c_idx, val in enumerate(row, start=7):  # column G
            ws.cell(row=r_idx, column=c_idx, value=val)

    # Donut chart
    fig, ax = plt.subplots(figsize=(6, 4))
    ax.pie(
        triage_summary.values,
        labels=triage_summary.index.tolist(),
        autopct=lambda p: f"{p:.1f}%" if p > 0 else "",
        startangle=90,
        wedgeprops=dict(width=0.45)
    )
    ax.set_title("Triage Categorisation (Proportion of Pairs)")
    plt.tight_layout()

    img_path = os.path.join(out_dir, "m4_donut.png")
    save_fig_to_png(fig, img_path)
    ws.add_image(XLImage(img_path), "G6")

    return triage_out, triage_summary, img_path
