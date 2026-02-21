# modules/M2b_entity_clustering.py
import os
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils.dataframe import dataframe_to_rows

from modules.helpers import find_col, write_df_to_sheet, save_fig_to_png

def run(src_path: str, wb_out, sheet_dup: str, sheet_input: str = None,
        tab_name: str = "M2b_EntityClustering", out_dir: str = "."):
    """
    Entity clustering (long tail + top IDs + pareto chart).
    Uses Input_DuplicatePatientSummary if present; else DuplicatePatientsSummary.
    Writes to Excel tab: M2b_EntityClustering
    Returns: long_tail_df, high_priority_df, img_path, cluster_source
    """

    cluster_source = sheet_input if sheet_input else sheet_dup

    cols = pd.read_excel(src_path, sheet_name=cluster_source, nrows=0, engine="openpyxl").columns.tolist()
    p1 = find_col(cols, ["Patient 1", "Patient1"])
    p2 = find_col(cols, ["Patient 2", "Patient2"])

    if p1 is None or p2 is None:
        # fallback to duplicate sheet
        cluster_source = sheet_dup
        cols = pd.read_excel(src_path, sheet_name=cluster_source, nrows=0, engine="openpyxl").columns.tolist()
        p1 = find_col(cols, ["Patient 1", "Patient1"])
        p2 = find_col(cols, ["Patient 2", "Patient2"])

    if p1 is None or p2 is None:
        raise ValueError("Could not find Patient 1 / Patient 2 columns for clustering.")

    df_cluster = pd.read_excel(src_path, sheet_name=cluster_source, usecols=[p1, p2], engine="openpyxl")

    ids = pd.concat([df_cluster[p1], df_cluster[p2]], ignore_index=True).dropna().astype(str)
    id_counts = ids.value_counts()

    freq_table = id_counts.reset_index()
    freq_table.columns = ["PatientID", "Associations"]

    def bucket(n):
        if n == 2: return "2"
        if n == 3: return "3"
        if n == 4: return "4"
        if n >= 5: return "5+"
        return "1"

    long_tail = id_counts.map(bucket).value_counts().reindex(["2", "3", "4", "5+"], fill_value=0)
    long_tail_df = pd.DataFrame({"Associations (pairs per ID)": long_tail.index, "Number of Patients": long_tail.values})

    high_priority = freq_table[freq_table["Associations"] >= 3].sort_values("Associations", ascending=False).head(20)

    sorted_counts = id_counts.sort_values(ascending=False).values
    fig, ax = plt.subplots(figsize=(7, 4))
    ax.plot(range(1, len(sorted_counts) + 1), sorted_counts, marker="o", markersize=2, linewidth=1)
    ax.axhline(2.5, color="red", linestyle="--", linewidth=1)
    ax.text(1, 2.55, "≥3 associations (high priority)", color="red", fontsize=8, va="bottom")
    ax.set_xlabel("Patient rank (highest to lowest associations)")
    ax.set_ylabel("Associations (appearances across Patient 1 & Patient 2)")
    ax.set_title("Entity Clustering Pareto (Association Distribution)")
    plt.tight_layout()

    img_path = os.path.join(out_dir, "m2b_entity_pareto.png")
    save_fig_to_png(fig, img_path)

    if tab_name in wb_out.sheetnames:
        ws = wb_out[tab_name]
    else:
        ws = wb_out.create_sheet(tab_name)

    ws.delete_rows(1, ws.max_row)
    ws.cell(row=1, column=1, value=f"Source sheet used: {cluster_source}")
    ws.cell(row=2, column=1, value="Long tail distribution (IDs with 2, 3, 4, 5+ associations):")

    for r_idx, row in enumerate(dataframe_to_rows(long_tail_df, index=False, header=True), start=3):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    start_row_hp = 3 + len(long_tail_df) + 3
    ws.cell(row=start_row_hp, column=1, value="High-priority Patient IDs (≥3 associations). Top 20 shown:")

    for r_idx, row in enumerate(dataframe_to_rows(high_priority, index=False, header=True), start=start_row_hp + 1):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    ws.add_image(XLImage(img_path), "E2")
    return long_tail_df, high_priority, img_path, cluster_source
