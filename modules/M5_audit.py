# modules/M5_audit.py
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows

from modules.helpers import write_df_to_sheet

def run(wb_out, audit_base: dict, audit_extra: dict = None, tab_name: str = "M5_Audit"):
    """
    Writes audit metadata to M5_Audit.
    - audit_base: key info (run date, source file, record counts, schema failures)
    - audit_extra: optional extra metrics appended underneath
    """
    if tab_name in wb_out.sheetnames:
        ws = wb_out[tab_name]
    else:
        ws = wb_out.create_sheet(tab_name)

    base_df = pd.DataFrame(list(audit_base.items()), columns=["Field", "Value"])
    write_df_to_sheet(ws, base_df, index=False)

    if audit_extra:
        extra_df = pd.DataFrame(list(audit_extra.items()), columns=["Field", "Value"])
        start_r = ws.max_row + 2
        for r_idx, row in enumerate(dataframe_to_rows(extra_df, index=False, header=True), start=start_r):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)

    return True
