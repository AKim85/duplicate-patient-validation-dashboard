# modules/M3_mismatch_patterns.py
import os
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image as XLImage

from modules.helpers import write_df_to_sheet, save_fig_to_png

def run(df_std: pd.DataFrame, wb_out, tab_name: str = "M3_MismatchPatterns", out_dir: str = "."):
    """
    Aggregate mismatch counts by MatchScore band (4–6).
    IMPORTANT: No patient IDs are written (aggregate-only).
    Expects df_std to contain:
      MatchScore_std and *_is_match boolean columns for each field.
    Writes to Excel tab: M3_MismatchPatterns
    Returns: mismatch_table, img_path
    """
    bands = [4, 5, 6]
    mismatch_fields = ["Forename", "Surname", "Soundex", "Sex", "DOB", "PostCode/Poscode", "Address"]

    mismatch_table = pd.DataFrame(index=bands, columns=mismatch_fields, dtype=int)

    for b in bands:
        sub = df_std[df_std["MatchScore_std"] == b]
        for fld in mismatch_fields:
            colm = f"{fld}_is_match"
            mismatch_table.loc[b, fld] = int((sub[colm] == False).sum()) if colm in sub.columns else 0

    mismatch_table = mismatch_table.fillna(0).astype(int)
    mismatch_table.index.name = "MatchScore"

    if tab_name in wb_out.sheetnames:
        ws = wb_out[tab_name]
    else:
        ws = wb_out.create_sheet(tab_name)

    write_df_to_sheet(ws, mismatch_table.reset_index(), index=False)
    ws.cell(row=mismatch_table.shape[0] + 3, column=1, value="Note: Output is aggregate-only and contains no patient IDs.")

    fig, ax = plt.subplots(figsize=(7, 4))
    mismatch_table.plot(kind="bar", stacked=True, ax=ax, colormap="tab20")
    ax.set_xlabel("MatchScore band")
    ax.set_ylabel("Number of mismatches (aggregate)")
    ax.set_title("Mismatch Patterns by MatchScore (4–6)")
    ax.legend(title="Field", bbox_to_anchor=(1.02, 1), loc="upper left", fontsize=7)
    plt.tight_layout()

    img_path = os.path.join(out_dir, "m3_mismatch_stacked.png")
    save_fig_to_png(fig, img_path)
    ws.add_image(XLImage(img_path), "J2")

    return mismatch_table, img_path

